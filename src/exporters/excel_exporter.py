"""Excel 导出器：4-Sheet（主表 + 子系表 + 组件 Sheet + 模块 Sheet）。"""
import os
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.models.sys_domain_change_usecase import SysDomainChangeBundle
from src.utils.logger import get_logger
from src.utils.text_clean import sanitize_for_excel

log = get_logger(__name__)


class ExcelExporter:
    """生成 4-Sheet 的 Excel。"""

    MAIN_HEADERS = [
        "rdc_id", "sub_system",
        "方案", "特性域变更分析结果",
        "需求描述", "算法描述", "实现思路",
    ]
    SUBSYSTEM_HEADERS = ["rdc_id", "sub_system", "change_points"]
    COMPONENT_HEADERS = [
        "rdc_id", "sub_system", "business_flow",
        "source_system", "aim_system",
        "change_type", "change_describe",
    ]
    MODULE_HEADERS = [
        "rdc_id", "sub_system", "sender", "receiver",
        "business_flow", "output_strategy", "change_type", "remarks",
    ]

    HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")

    DEFAULT_DIR = "output"
    DEFAULT_FILENAME = "sys_domain_change_usecase.xlsx"

    def __init__(self, output_dir: str = DEFAULT_DIR):
        self.output_dir = output_dir

    def export(self, bundles: List[SysDomainChangeBundle]) -> str:
        os.makedirs(self.output_dir, exist_ok=True)
        path = os.path.join(self.output_dir, self.DEFAULT_FILENAME)

        wb = Workbook()
        wb.remove(wb.active)

        self._write_main_sheet(wb, bundles)
        self._write_subsystem_sheet(wb, bundles)
        self._write_component_sheet(wb, bundles)
        self._write_module_sheet(wb, bundles)

        wb.save(path)
        log.info("Excel 已写入: {}", path)
        return path

    # ---------- Sheet 1: 主表（1 行/rdc_id） ----------
    def _write_main_sheet(self, wb: Workbook, bundles: List[SysDomainChangeBundle]):
        ws = wb.create_sheet("主表")
        self._write_header(ws, self.MAIN_HEADERS)

        for b in bundles:
            if b.main is None:
                continue
            m = b.main
            row = [
                m.rdc_id,
                m.sub_system,            # joined sub_system 字符串
                m.solution,
                m.feature_change_analysis,
                m.requirement,
                m.algorithm,
                m.implementation,
            ]
            self._append_row(ws, row)

        widths = [16, 18, 50, 40, 40, 40, 40]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

    # ---------- Sheet 2: 子系表 ----------
    def _write_subsystem_sheet(self, wb: Workbook, bundles: List[SysDomainChangeBundle]):
        ws = wb.create_sheet("子系表")
        self._write_header(ws, self.SUBSYSTEM_HEADERS)

        for b in bundles:
            for s in b.subsystems:
                cps = "\n".join(s.change_points) if s.change_points else ""
                self._append_row(ws, [s.rdc_id, s.sub_system, cps])

        widths = [16, 18, 50]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

    # ---------- Sheet 3: 组件 ----------
    def _write_component_sheet(self, wb: Workbook, bundles: List[SysDomainChangeBundle]):
        ws = wb.create_sheet("波及组件")
        self._write_header(ws, self.COMPONENT_HEADERS)

        for b in bundles:
            for c in b.components:
                self._append_row(ws, [
                    c.rdc_id, c.sub_system, c.business_flow,
                    c.source_system, c.aim_system,
                    c.change_type, c.change_describe,
                ])

        widths = [16, 18, 22, 22, 22, 22, 30]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

    # ---------- Sheet 4: 模块 ----------
    def _write_module_sheet(self, wb: Workbook, bundles: List[SysDomainChangeBundle]):
        ws = wb.create_sheet("波及模块")
        self._write_header(ws, self.MODULE_HEADERS)

        for b in bundles:
            for mod in b.modules:
                self._append_row(ws, [
                    mod.rdc_id, mod.sub_system,
                    mod.sender, mod.receiver,
                    mod.business_flow, mod.output_strategy,
                    mod.change_type, mod.remarks,
                ])

        widths = [16, 18, 18, 18, 30, 22, 22, 30]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

    # ---------- 工具 ----------
    def _write_header(self, ws, headers: List[str]):
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _append_row(self, ws, row: list):
        clean_row = [sanitize_for_excel(v) for v in row]
        ws.append(clean_row)
        for cell in ws[ws.max_row]:
            cell.alignment = self.WRAP_ALIGN
